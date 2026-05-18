"""Generate blank Excel templates for data onboarding.

Usage:
    python data/onboarding/templates/generate_templates.py [output_dir]

Outputs:
    output_dir/admission_template.xlsx
    output_dir/curriculum_template.xlsx
    output_dir/employment_template.xlsx
"""
import os
import sys

import openpyxl


def generate_admission_template(output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "录取分数线"
    ws.append(["year", "province", "batch", "major_name", "min_score",
               "min_rank", "subject_requirements", "enrollment_quota"])
    ws.append([2025, "广东", "本科批", "计算机科学与技术", 589, 28500, "物理", 120])
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    wb.save(output_path)


def generate_curriculum_template(output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "培养计划"
    ws.append(["major_name", "college", "duration", "core_courses",
               "objective", "degree"])
    ws.append(["计算机科学与技术", "计算机学院", "4年",
               "数据结构,操作系统,计算机组成原理,计算机网络,算法设计",
               "培养具备计算机系统设计与开发能力的高级工程技术人才", "工学学士"])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 12
    wb.save(output_path)


def generate_employment_template(output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "就业报告"
    ws.append(["major_name", "year", "employment_rate", "avg_salary",
               "main_industries", "typical_companies", "further_study_rate"])
    ws.append(["计算机科学与技术", 2024, 96.5, 15000,
               "互联网,金融,教育", "腾讯,阿里巴巴,华为,字节跳动", 15.0])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 14
    wb.save(output_path)


TEMPLATES = {
    "admission_template.xlsx": generate_admission_template,
    "curriculum_template.xlsx": generate_curriculum_template,
    "employment_template.xlsx": generate_employment_template,
}


if __name__ == "__main__":
    output_dir = sys.argv[1] if len(sys.argv) > 1 else os.path.dirname(__file__)
    os.makedirs(output_dir, exist_ok=True)
    for filename, generator in TEMPLATES.items():
        path = os.path.join(output_dir, filename)
        generator(path)
        print(f"Generated: {path}")
