import uuid
from datetime import datetime, timezone

import openpyxl

from models import async_session
from tenants.models import TenantData
from knowledge.indexer import index_tenant_data
from data.onboarding.validator import validate_admission_row, validate_curriculum_row, validate_employment_row

IMPORT_CONFIG = {
    "admission_score": (validate_admission_row, [
        "year", "province", "batch", "major_name", "min_score",
        "min_rank", "subject_requirements", "enrollment_quota"
    ]),
    "curriculum": (validate_curriculum_row, [
        "major_name", "college", "duration", "core_courses", "objective", "degree"
    ]),
    "employment": (validate_employment_row, [
        "major_name", "year", "employment_rate", "avg_salary",
        "main_industries", "typical_companies", "further_study_rate"
    ]),
}


async def import_excel(
    tenant_id: str,
    tenant_slug: str,
    file_path: str,
    data_type: str,
) -> dict:
    validator, expected_columns = IMPORT_CONFIG[data_type]

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    missing = [c for c in expected_columns if c not in headers]
    if missing:
        return {"success": False, "error": f"缺少列: {missing}", "imported": 0}

    rows = []
    errors = []
    total_rows = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        total_rows += 1
        data = dict(zip(headers, row))
        err = validator(data)
        if err:
            errors.append(f"Row {i}: {err}")
        else:
            rows.append(data)

    seen = set()
    deduped = []
    for r in rows:
        key = (r.get("year"), r.get("province"), r.get("major_name"))
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    async with async_session() as db:
        for r in deduped:
            td = TenantData(
                id=uuid.uuid4(),
                tenant_id=tenant_id,
                data_type=data_type,
                title=f"{r.get('major_name', '')} {r.get('year', '')} {r.get('province', '')}",
                content=r,
                year=int(r.get("year", 0)) if r.get("year") else None,
                province=r.get("province", ""),
                extra_meta={
                    "major_name": r.get("major_name", ""),
                    "subject_requirements": r.get("subject_requirements", ""),
                },
            )
            db.add(td)
            try:
                await index_tenant_data(tenant_slug, td)
                td.indexed_at = datetime.now(timezone.utc)
            except Exception as e:
                errors.append(f"Index failed for {td.title}: {e}")

        await db.commit()

    return {
        "success": True,
        "imported": len(deduped),
        "duplicates_skipped": len(rows) - len(deduped),
        "errors": errors,
        "total_rows": total_rows,
    }
